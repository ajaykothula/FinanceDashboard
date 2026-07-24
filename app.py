from flask import Flask, render_template, request, redirect, send_file, session
import os
from openpyxl import Workbook, load_workbook
import sqlite3
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

app = Flask(__name__)
app.secret_key = "finance_dashboard_secret"
DB = "database/finance.db"
def get_conn():
    conn = sqlite3.connect(DB, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn
@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        conn = sqlite3.connect(DB, timeout=10)
        cur = conn.cursor()

        cur.execute(
            "SELECT * FROM users WHERE username=? AND password=?",
            (username, password)
        )

        user = cur.fetchone()
        conn.close()

        if user:
            session["user"] = username
            return redirect("/")

        return "❌ Invalid username or password"

    return render_template("login.html")

@app.route("/signup", methods=["GET", "POST"])
def signup():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        conn = sqlite3.connect(DB, timeout=10)
        cur = conn.cursor()

        try:
            cur.execute(
                "INSERT INTO users(username,password) VALUES(?,?)",
                (username, password)
            )
            conn.commit()
            conn.close()

            return redirect("/login")

        except sqlite3.IntegrityError:
            conn.close()
            return "Username already exists!"

    return render_template("signup.html")
@app.route("/logout")
def logout():

    session.clear()

    return redirect("/login")
def get_conn():
    return sqlite3.connect(DB, timeout=10)

@app.route("/")
def home():
    if "user" not in session:
        return redirect("/login")

    conn = sqlite3.connect(DB, timeout=10)
    cur = conn.cursor()

    cur.execute("SELECT IFNULL(SUM(amount),0) FROM transactions WHERE type='Income'")
    income = cur.fetchone()[0]

    cur.execute("SELECT IFNULL(SUM(amount),0) FROM transactions WHERE type='Expense'")
    expense = cur.fetchone()[0]

    balance = income - expense
    if income > 0:
       savings_rate = (balance / income) * 100
    else:
       savings_rate = 0
    cur.execute("SELECT goal_amount FROM savings_goal LIMIT 1")
    row = cur.fetchone()

    if row:
       goal = row[0]
    else:
       goal = 0

    if goal > 0:
       goal_percent = (balance / goal) * 100
       remaining_goal = goal - balance
    else:
       goal_percent = 0
       remaining_goal = 0
      

    cur.execute("SELECT COUNT(*) FROM transactions")
    count = cur.fetchone()[0]


    # ⭐ ADD BUDGET CODE HERE

    cur.execute("SELECT monthly_budget FROM budget LIMIT 1")
    row = cur.fetchone()

    if row:
        monthly_budget = row[0]
    else:
        monthly_budget = 0


# Get current month expenses only
    current_month = datetime.now().strftime("%Y-%m")

    cur.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM transactions
        WHERE type='Expense'
        AND date LIKE ?
     """, (current_month + "%",))

    monthly_expense = cur.fetchone()[0]


    remaining = monthly_budget - monthly_expense

    if monthly_budget > 0:
       budget_percent = (monthly_expense / monthly_budget) * 100
    else:
       budget_percent = 0

    # Get transactions
    # Search transactions
    # Search + Date Filter
    search = request.args.get("search", "")
    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")

    query = "SELECT * FROM transactions WHERE 1=1"
    params = []

    if search:
       query += " AND category LIKE ?"
       params.append("%" + search + "%")

    if from_date:
       query += " AND date >= ?"
       params.append(from_date)

    if to_date:
       query += " AND date <= ?"
       params.append(to_date)
 
    query += " ORDER BY id"

    cur.execute(query, params)
    transactions = cur.fetchall()
    # Monthly chart data
    # Monthly chart data
    cur.execute("""
        SELECT
            strftime('%Y-%m', date),
            SUM(CASE WHEN type='Income' THEN amount ELSE 0 END),
            SUM(CASE WHEN type='Expense' THEN amount ELSE 0 END)
        FROM transactions
        GROUP BY strftime('%Y-%m', date)
        ORDER BY strftime('%Y-%m', date)
    """)

    monthly_data = cur.fetchall()
    cur.execute("""
         SELECT category, SUM(amount)
           FROM transactions
             WHERE type='Expense'
           GROUP BY category
          """)

    chart_data = cur.fetchall()
    cur.execute("""
    SELECT category, SUM(amount)
    FROM transactions
    WHERE type='Expense'
    GROUP BY category
    ORDER BY SUM(amount) DESC
    """)

    category_summary = cur.fetchall()
    conn.close()

    return render_template(
        "index.html",
        transactions=transactions,
        income=income,
        expense=expense,
        balance=balance,
        count=count,
        monthly_budget=monthly_budget,
        remaining=remaining,
        budget_percent=budget_percent,
        monthly_data=monthly_data,
        savings_rate=savings_rate,
        chart_data=chart_data,
        category_summary=category_summary,
        goal=goal,
        goal_percent=goal_percent,
        remaining_goal=remaining_goal
    )
@app.route("/add", methods=["POST"])
def add():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO transactions(date,category,type,amount,description)
        VALUES(?,?,?,?,?)
    """, (
        request.form["date"],
        request.form["category"],
        request.form["type"],
        request.form["amount"],
        request.form["description"]
    ))

    conn.commit()
    conn.close()

    return redirect("/")

@app.route("/delete/<int:id>")
def delete(id):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("DELETE FROM transactions WHERE id=?", (id,))

    conn.commit()
    conn.close()

    return redirect("/")

@app.route("/edit/<int:id>", methods=["GET", "POST"])
def edit(id):
    conn = get_conn()
    cur = conn.cursor()

    if request.method == "POST":
        cur.execute("""
            UPDATE transactions
            SET date=?,category=?,type=?,amount=?,description=?
            WHERE id=?
        """, (
            request.form["date"],
            request.form["category"],
            request.form["type"],
            request.form["amount"],
            request.form["description"],
            id
        ))

        conn.commit()
        conn.close()
        return redirect("/")

    cur.execute("SELECT * FROM transactions WHERE id=?", (id,))
    transaction = cur.fetchone()

    conn.close()

    return render_template("edit.html", transaction=transaction)
@app.route("/transactions")
def all_transactions():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, date, category, type, amount, description
        FROM transactions
        ORDER BY id DESC
    """)

    transactions = cur.fetchall()

    conn.close()

    return render_template(
        "transactions.html",
        transactions=transactions
    )
@app.route("/export")
def export():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM transactions")
    rows = cur.fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws.append([
        "ID",
        "Date",
        "Category",
        "Type",
        "Amount",
        "Description"
    ])
    for row in rows:
        ws.append(row)

    filename = "transactions.xlsx"

    wb.save(filename)

    return send_file(
        os.path.abspath(filename),
        as_attachment=True,
        download_name="transactions.xlsx"
)
@app.route("/set_budget", methods=["POST"])
def set_budget():

    budget = request.form["budget"]

    conn = sqlite3.connect(DB, timeout=10)
    cur = conn.cursor()

    cur.execute("DELETE FROM budget")

    cur.execute(
        "INSERT INTO budget(monthly_budget) VALUES(?)",
        (budget,)
    )

    conn.commit()
    conn.close()

    return redirect("/")
@app.route("/save_goal", methods=["POST"])
def save_goal():

    conn = get_conn()
    cur = conn.cursor()

    goal = request.form["goal"]

    cur.execute("DELETE FROM savings_goal")

    cur.execute(
        "INSERT INTO savings_goal(goal_amount) VALUES(?)",
        (goal,)
    )

    conn.commit()
    conn.close()

    return redirect("/")
@app.route("/reset_goal")
def reset_goal():

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("DELETE FROM savings_goal")

    conn.commit()
    conn.close()

    return redirect("/")
@app.route("/reset_budget")
def reset_budget():

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("DELETE FROM budget")

    conn.commit()
    conn.close()

    return redirect("/")

@app.route("/import", methods=["POST"])
def import_excel():
    file = request.files["file"]

    wb = load_workbook(file)
    ws = wb.active

    # Read the first row (headers)
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    # Create a mapping of header name -> column index
    cols = {name: i for i, name in enumerate(headers)}

    required = ["date", "category", "type", "amount", "description"]

    # Check required columns exist
    for col in required:
        if col not in cols:
            return f"❌ Missing required column: {col}"

    conn = get_conn()
    cur = conn.cursor()

    for row in ws.iter_rows(min_row=2, values_only=True):

        date = row[cols["date"]]
        category = row[cols["category"]]
        ttype = row[cols["type"]]
        amount = row[cols["amount"]]
        description = row[cols["description"]]

        cur.execute("""
            INSERT INTO transactions
            (date, category, type, amount, description)
            VALUES (?, ?, ?, ?, ?)
        """, (date, category, ttype, amount, description))

    conn.commit()
    conn.close()

    return redirect("/")
@app.route("/export_pdf")
def export_pdf():

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT IFNULL(SUM(amount),0) FROM transactions WHERE type='Income'")
    income = cur.fetchone()[0]

    cur.execute("SELECT IFNULL(SUM(amount),0) FROM transactions WHERE type='Expense'")
    expense = cur.fetchone()[0]

    balance = income - expense

    cur.execute("SELECT COUNT(*) FROM transactions")
    count = cur.fetchone()[0]

    conn.close()

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate("finance_report.pdf")
    story = []

    story.append(Paragraph("<b>Mini Finance Tracker Report</b>", styles["Title"]))
    story.append(Paragraph(f"Total Income: ₹{income}", styles["Normal"]))
    story.append(Paragraph(f"Total Expense: ₹{expense}", styles["Normal"]))
    story.append(Paragraph(f"Balance: ₹{balance}", styles["Normal"]))
    story.append(Paragraph(f"Transactions: {count}", styles["Normal"]))

    doc.build(story)

    return send_file(
        "finance_report.pdf",
        as_attachment=True
    )
@app.route("/monthly_report")
def monthly_report():

    if "user" not in session:
        return redirect("/login")

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            substr(date,1,7) AS month,
            SUM(CASE WHEN type='Income' THEN amount ELSE 0 END) AS income,
            SUM(CASE WHEN type='Expense' THEN amount ELSE 0 END) AS expense,
            COUNT(*) AS transactions
        FROM transactions
        GROUP BY substr(date,1,7)
        ORDER BY month DESC
    """)

    reports = cur.fetchall()

    conn.close()

    return render_template(
        "monthly_report.html",
        reports=reports
    )
@app.route("/clear_transactions")
def clear_transactions():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    cur.execute("DELETE FROM transactions")
    cur.execute("DELETE FROM sqlite_sequence WHERE name='transactions'")

    conn.commit()
    conn.close()

    return redirect("/")
@app.route("/dynamic_import", methods=["POST"])
def dynamic_import():

    file = request.files["file"]

    wb = load_workbook(file)

    ws = wb.active

    headers = [str(cell.value).strip() for cell in ws[1]]

    conn = get_conn()
    cur = conn.cursor()

    table_name = "uploaded_data"

    columns = []

    for h in headers:
        col = h.replace(" ", "_").replace("-", "_")
        columns.append(f'"{col}" TEXT')

    sql = f'''
    CREATE TABLE IF NOT EXISTS {table_name} (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        {",".join(columns)}
    )
    '''

    print(sql)

    cur.execute(sql)

# Insert all rows
    for row in ws.iter_rows(min_row=2, values_only=True):

        placeholders = ",".join(["?"] * len(headers))

        column_names = ",".join([f'"{h}"' for h in headers])

        cur.execute(
                     f'''
               INSERT INTO uploaded_data ({column_names})
               VALUES ({placeholders})
               ''',
                row
    )

    conn.commit()
    conn.close()

    return redirect("/data-manager")
@app.route("/data_manager")
def data_manager():
    return render_template("data_manager.html")
@app.route("/profile")
def profile():

    if "user" not in session:
        return redirect("/login")

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT username 
    FROM users 
    WHERE username=?
    """,(session["user"],))

    user = cur.fetchone()

    cur.execute("""
    SELECT 
    SUM(CASE WHEN type='Income' THEN amount ELSE 0 END),
    SUM(CASE WHEN type='Expense' THEN amount ELSE 0 END)
    FROM transactions
    """)

    data = cur.fetchone()

    conn.close()

    income = data[0] or 0
    expense = data[1] or 0
    balance = income - expense

    return render_template(
        "profile.html",
        username=user[0],
        income=income,
        expense=expense,
        balance=balance
    )

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
